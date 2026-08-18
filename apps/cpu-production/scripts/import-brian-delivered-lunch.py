"""Convert Brian's weekly workbook into source-backed delivered-in lunch fixtures.

This is deliberately a conversion step, not a publisher.  It preserves the
workbook wording, allergen evidence and raw site columns so an operator can
review mappings before any canonical/site records are created.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

PARENT = "delivered-in-lunch"
ALLERGEN_KEYS = [
    "noKeyAllergens", "peanuts", "otherNuts", "gluten", "sesame", "molluscs",
    "fish", "soya", "celery", "shellfish", "eggs", "milk", "mustard",
    "lupin", "sulphites",
]
DAY_SHEETS = {"mon": "Monday", "tue": "Tuesday", "wed": "Wednesday", "thurs": "Thursday", "fri": "Friday"}
SITE_LABELS = {
    "angel": "Angel Court", "angeel": "Angel Court", "angel court": "Angel Court",
    "bp": "Bridgepoint", "bridgepoint": "Bridgepoint",
    "mk": "MNK", "mnk": "MNK", "comm": "Commerzbank", "commerzbank": "Commerzbank",
    # Haleon is intentionally source-only until an OPLOC is governed.
    "haelon": "Haleon (source-only)",
    "x": "Unresolved source column X",
}


def clean(value):
    return " ".join(str(value or "").split()).strip()


def slug(value):
    return re.sub(r"[^a-z0-9]+", "-", clean(value).lower()).strip("-") or "untitled"


def marker(value):
    text = clean(value).lower()
    if text in {"x", "yes", "contains", "1"}:
        return "contains"
    if text in {"mc", "m/c", "may contain", "may"}:
        return "may_contain"
    return ""


def production_rows(ws, source_name, day):
    rows = []
    headers = [clean(v) for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    if len(headers) < 2:
        return rows
    total_index = next((i for i, h in enumerate(headers[2:], 2) if h.lower() == "total"), len(headers))
    site_columns = [(i, h) for i, h in enumerate(headers[2:total_index], 2) if h]
    for row_number, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        if not row:
            continue
        product, dish = clean(row[0]), clean(row[1] if len(row) > 1 else "")
        if not dish or dish.lower() in {"dish", "date", "verified by location manager"}:
            continue
        quantities = []
        for index, raw_key in site_columns:
            value = row[index] if index < len(row) else None
            try:
                quantity = float(value) if value not in (None, "") else 0
            except (TypeError, ValueError):
                quantity = 0
            if quantity <= 0:
                continue
            key = clean(raw_key).lower()
            quantities.append({
                "sourceSiteKey": key,
                "siteLabel": SITE_LABELS.get(key, clean(raw_key)),
                "canonicalOplocId": None,
                "quantity": int(quantity) if quantity.is_integer() else quantity,
                "mappingStatus": "source-only" if key not in SITE_LABELS or key == "haelon" else "label-confirmed-id-pending",
            })
        rows.append({"day": day, "sourceSheet": source_name, "sourceRow": row_number, "productGroup": product, "title": dish, "siteQuantities": quantities})
    return rows


def allergen_rows(ws, source_name, day):
    header_row = None
    for number, row in enumerate(ws.iter_rows(values_only=True), 1):
        if clean(row[0]).lower() == "dish / food / product":
            header_row = number
            break
    if not header_row:
        return []
    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        title = clean(row[0] if row else "")
        if not title or title.lower() in {"verified by location manager:", "week ending date:"}:
            continue
        if title.lower() in {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}:
            continue
        values = OrderedDict()
        for offset, key in enumerate(ALLERGEN_KEYS):
            values[key] = marker(row[3 + offset] if 3 + offset < len(row) else None)
        note = clean(row[18] if len(row) > 18 else "")
        rows.append({"day": day, "sourceSheet": source_name, "sourceRow": row_number, "title": title, "allergens": values, "mayContainNotes": note})
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parents[1] / "data")
    args = parser.parse_args()
    wb = load_workbook(args.source, data_only=True, read_only=True)
    plans, evidence = [], []
    for sheet, day in DAY_SHEETS.items():
        if sheet in wb.sheetnames:
            plans.extend(production_rows(wb[sheet], sheet, day))
        allergen_sheet = f"fika{sheet}"
        if allergen_sheet in wb.sheetnames:
            evidence.extend(allergen_rows(wb[allergen_sheet], allergen_sheet, day))

    by_key = OrderedDict()
    for record in evidence:
        key = slug(record["title"])
        item = by_key.setdefault(key, {
            "id": f"production:delivered-in-lunch:{key}",
            "title": record["title"],
            "allergens": dict(record["allergens"]),
            "mayContainNotes": record["mayContainNotes"],
            "itemType": "other",
            "parentMenuItemKey": PARENT,
            "sourceEvidence": [],
        })
        item["sourceEvidence"].append(f"{record['sourceSheet']}!A{record['sourceRow']}")
        if record["mayContainNotes"] and record["mayContainNotes"] not in item["mayContainNotes"]:
            item["mayContainNotes"] = "; ".join(filter(None, [item["mayContainNotes"], record["mayContainNotes"]]))
        for allergen, state in record["allergens"].items():
            if state == "contains" or (state == "may_contain" and not item["allergens"].get(allergen)):
                item["allergens"][allergen] = state

    # Keep generated fixtures reproducible across reruns; the workbook week is
    # the source version, not the machine's current clock.
    generated = "2026-07-20T00:00:00Z"
    for item in by_key.values():
        item["sourceEvidence"] = sorted(set(item["sourceEvidence"]))
        item["updatedAt"] = generated
        item["updatedBy"] = "brian-workbook-import"
    plan = {
        "schemaVersion": "0.1.0",
        "parentMenuItemKey": PARENT,
        "sourceFile": args.source.name,
        "sourceWeekCommencing": "2026-07-20",
        "mappingPolicy": "Site labels are preserved as evidence; canonical OPLOC IDs are intentionally unset until governed.",
        "plans": plans,
        "allergenEvidence": evidence,
        "items": list(by_key.values()),
    }
    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "delivered-in-lunch-plan-2026-07-20.json").write_text(json.dumps(plan, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    (args.output_dir / "delivered-in-lunch-items-seed.json").write_text(json.dumps(list(by_key.values()), indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"items": len(by_key), "productionRows": len(plans), "allergenRows": len(evidence), "outputDir": str(args.output_dir)}))


if __name__ == "__main__":
    main()
